import argparse
import re

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles import Alignment, Font, Border, PatternFill


def get_or_create_sheet(book, sheet_name):
    if sheet_name in book.sheetnames:
        return book[sheet_name]
    return book.create_sheet(title=sheet_name)


def set_column_and_row_formatting(sheet):
    # A:AB  ->  1..28 (A=1, ..., Z=26, AA=27, AB=28)
    for col_idx in range(1, 29):
        col_letter = get_column_letter(col_idx)
        sheet.column_dimensions[col_letter].width = 8.6

    first_row = sheet[1]  # row 1 (tuple of cells)
    for cell in first_row:
        cell.alignment = Alignment(
            wrap_text=True,
            horizontal="left",
            vertical="bottom",
        )
        cell.font = Font(bold=False)
        cell.border = Border()


def generate_column_names(headers):
    names = {}
    for idx, key in enumerate(headers, start=1):
        names[key] = get_column_letter(idx)
    return names


def insert_headers(sheet, headers):
    for key, col_letter in headers.items():
        cell = sheet[f"{col_letter}1"]
        if cell.value != key:
            idx = column_index_from_string(col_letter)
            sheet.insert_cols(idx)
            sheet[f"{col_letter}1"] = key


def find_last_row(sheet, column_letter="B"):
    col_idx = column_index_from_string(column_letter)
    last_data_row = None
    for row in range(sheet.max_row, 1, -1):
        value = sheet.cell(row=row, column=col_idx).value
        if value not in (None, ""):
            last_data_row = row
            break
    return last_data_row or 2


CELL_REF_ROW2_RE = re.compile(r"(?<!\$)([A-Z]{1,3})2(?!\d)")


def shift_formula_from_row2(formula: str, target_row: int) -> str:
    if target_row == 2 or not isinstance(formula, str) or not formula.startswith("="):
        return formula

    def repl(match):
        col = match.group(1)
        return f"{col}{target_row}"

    return CELL_REF_ROW2_RE.sub(repl, formula)


def apply_formulas(sheet, formulas, headers):
    for key, formula in formulas.items():
        col_letter = headers[key]
        sheet[f"{col_letter}2"].value = formula


def autofill_formulas(sheet, formulas, headers):
    last_row = find_last_row(sheet, "B")
    if last_row <= 2:
        return

    for key in formulas.keys():
        col_letter = headers[key]
        base_cell = sheet[f"{col_letter}2"]
        base_formula = base_cell.value
        if not isinstance(base_formula, str) or not base_formula.startswith("="):
            continue

        for row in range(3, last_row + 1):
            sheet[f"{col_letter}{row}"].value = shift_formula_from_row2(base_formula, row)


def hide_columns_range(sheet, start_col_letter, end_col_letter):
    start_idx = column_index_from_string(start_col_letter)
    end_idx = column_index_from_string(end_col_letter)
    for idx in range(start_idx, end_idx + 1):
        col_letter = get_column_letter(idx)
        sheet.column_dimensions[col_letter].hidden = True


def color_entire_column(sheet, col_letter, rgb_tuple):
    r, g, b = rgb_tuple
    hex_color = f"{r:02X}{g:02X}{b:02X}"
    fill = PatternFill(fill_type="solid", start_color="FF" + hex_color, end_color="FF" + hex_color)

    for row in range(1, sheet.max_row + 1):
        cell = sheet[f"{col_letter}{row}"]
        cell.fill = fill


def main(filename=None, echo=True):
    parser = argparse.ArgumentParser()
    parser.add_argument("--excelname", help="Excel file name")
    args = parser.parse_args()
    filename = filename or args.excelname

    if not filename:
        raise SystemExit("Excel filename is required (either argument or --excelname)")

    book = load_workbook(filename, data_only=False)

    sheet1 = book["Sheet1"]
    sheet3 = book["Sheet3"]

    set_column_and_row_formatting(sheet1)

    sheet1.freeze_panes = "A2"

    sheet3.column_dimensions["A"].width = 15
    set_column_and_row_formatting(sheet3)

    max_days = round(book["Sheet4"]["A2"].value)
    can_be_sold_total = f"Can be sold in {max_days} days"

    headers_sheet1 = [
        "Deal ID",
        "Item No",
        "Item Name",
        "Used Minimum Order Quantity",
        "Deal Sum",
        "Purchase Price",
        "Sale Price",
        "Profit",
        "Average Daily Sales",
        "Empty ADS Deal",
        "Inventory",
        "Can be sold in credit terms",
        can_be_sold_total,
        "System Suggested Quantity",
        "Best suggested quantity",
        "Overstock",
        "Days For Sale",
        "Deal Days Dispersion",
        "Item Budget",
        "Budget",
        "Total Item Sales",
        "Total Item Profit",
        "30 Days Profit",
        "30 Days Sales",
        "Weighted Average 30 Day Profit Margin",
        "Effectiveness",
        "Average Effectiveness",
        "AddToTotal",
        "Included",
        "Days For Sale Average",
        "Days For Sale sub",
        "Overstock check",
        "Overstock check sum",
        "Daily Sales",
        "Empty ADS Check Profit",
    ]

    headers_sheet3 = ["Deal ID", "ABC", "Daily Sales", "MaxBQ", "BSQ", "SSQ"]

    names1 = generate_column_names(headers_sheet1)
    names3 = generate_column_names(headers_sheet3)

    insert_headers(sheet1, names1)
    insert_headers(sheet3, names3)

    formulas_sheet1 = {
        "Used Minimum Order Quantity": (
            f'=_xlfn.MAXIFS(Sheet2!D:D, Sheet2!B:B, {names1["Item No"]}2, '
            f'Sheet2!D:D, "<="&{names1["Deal Sum"]}2)'
        ),
        "Deal Sum": (
            f"=SUMIF({names1['Deal ID']}:{names1['Deal ID']}, {names1['Deal ID']}2, "
            f"{names1['Best suggested quantity']}:{names1['Best suggested quantity']})"
        ),
        "Empty ADS Deal": (
            f"=IF(AND(SUMIFS({names1['Average Daily Sales']}:{names1['Average Daily Sales']},"
            f"{names1['Deal ID']}:{names1['Deal ID']},{names1['Deal ID']}2)<=0,"
            f"COUNTIFS({names1['Deal ID']}:{names1['Deal ID']},{names1['Deal ID']}2,"
            f"{names1['Empty ADS Check Profit']}:{names1['Empty ADS Check Profit']},\"<0\")),\"empty\",\"OK\")"
        ),
        "Purchase Price": (
            f'=_xlfn.MINIFS(Sheet2!E:E, Sheet2!B:B, {names1["Item No"]}2, '
            f'Sheet2!D:D, "<="&{names1["Deal Sum"]}2)'
        ),
        "Profit": f"={names1['Sale Price']}2 - {names1['Purchase Price']}2",
        can_be_sold_total: (
            f"=MAX(0, {names1['Average Daily Sales']}2 * Sheet4!$A$2 - {names1['Inventory']}2)"
        ),
        "Overstock": f"=IF({names1['Overstock check sum']}2,\"check\",\"OK\")",
        "Days For Sale": (
            f"=IF({names1['Average Daily Sales']}2 > 0, "
            f"({names1['Inventory']}2 + {names1['Best suggested quantity']}2) / "
            f"{names1['Average Daily Sales']}2, 0)"
        ),
        "Deal Days Dispersion": (
            f"=IF(COUNTIFS({names1['Deal ID']}:{names1['Deal ID']}, {names1['Deal ID']}2, "
            f"{names1['Included']}:{names1['Included']}, TRUE) > 0, "
            f"AVERAGEIFS({names1['Days For Sale sub']}:{names1['Days For Sale sub']}, "
            f"{names1['Deal ID']}:{names1['Deal ID']}, {names1['Deal ID']}2, "
            f"{names1['Included']}:{names1['Included']}, TRUE) / "
            f"COUNTIFS({names1['Deal ID']}:{names1['Deal ID']}, {names1['Deal ID']}2, "
            f"{names1['Included']}:{names1['Included']}, TRUE), 0)"
        ),
        "Item Budget": f"={names1['Purchase Price']}2 * {names1['Best suggested quantity']}2",
        "Budget": f"=SUM({names1['Item Budget']}:{names1['Item Budget']})",
        "Total Item Sales": f"={names1['Sale Price']}2 * {names1['Best suggested quantity']}2",
        "Total Item Profit": f"={names1['Best suggested quantity']}2 * {names1['Profit']}2",
        "30 Days Profit": (
            f"={names1['Profit']}2 * MIN({names1['Best suggested quantity']}2, "
            f"MAX(30 * {names1['Average Daily Sales']}2 - {names1['Inventory']}2, 0))"
        ),
        "30 Days Sales": (
            f"={names1['Sale Price']}2 * MIN({names1['Best suggested quantity']}2, "
            f"MAX(30 * {names1['Average Daily Sales']}2 - {names1['Inventory']}2, 0))"
        ),
        "Weighted Average 30 Day Profit Margin": (
            f"={names1['30 Days Profit']}2 / SUM({names1['30 Days Sales']}:"
            f"{names1['30 Days Sales']})"
        ),
        "Effectiveness": f"={names1['30 Days Profit']}2 - {names1['Deal Days Dispersion']}2",
        "Average Effectiveness": f"=SUM({names1['Effectiveness']}:{names1['Effectiveness']})",
        "Included": (
            f"=IF({names1['Average Daily Sales']}2 > 0, "
            f"({names1['Inventory']}2 + {names1['System Suggested Quantity']}2) / "
            f"{names1['Average Daily Sales']}2 <= Sheet4!$A$2 + "
            f"{names1['AddToTotal']}2 / {names1['Average Daily Sales']}2, FALSE)"
        ),
        "Days For Sale Average": (
            f"=IF({names1['Included']}2, "
            f"AVERAGEIFS({names1['Days For Sale']}:{names1['Days For Sale']}, "
            f"{names1['Deal ID']}:{names1['Deal ID']}, {names1['Deal ID']}2, "
            f"{names1['Included']}:{names1['Included']}, TRUE), 0)"
        ),
        "Days For Sale sub": (
            f"=IF({names1['Average Daily Sales']}2 > 0, "
            f"({names1['Days For Sale']}2 - {names1['Days For Sale Average']}2)^2, 0)"
        ),
        "Overstock check": (
            f"=IF(AND({names1['Days For Sale']}2>Sheet4!$A$2, "
            f"{names1['Best suggested quantity']}2>0), "
            f"IF(AND({names1['System Suggested Quantity']}2=1, "
            f"{names1['Best suggested quantity']}2=1), 0, 1), 0)"
        ),
        "Overstock check sum": (
            f"=SUMIF({names1['Deal ID']}:{names1['Deal ID']}, {names1['Deal ID']}2, "
            f"{names1['Overstock check']}:{names1['Overstock check']})"
        ),
        "Daily Sales": f"={names1['Sale Price']}2 * {names1['Average Daily Sales']}2",
        "Empty ADS Check Profit": (
            f"={names1['Profit']}2 * {names1['Best suggested quantity']}2"
        ),
    }

    formulas_sheet3 = {
        "Daily Sales": (
            f"=SUMIF(Sheet1!{names1['Deal ID']}:"
            f"{names1['Deal ID']}, {names3['Deal ID']}2, "
            f"Sheet1!{names1['Daily Sales']}:"
            f"{names1['Daily Sales']})"
        ),
        "MaxBQ": f"=_xlfn.MAXIFS(Sheet2!D:D, Sheet2!A:A, {names3['Deal ID']}2)",
        "BSQ": (
            f"=SUMIF(Sheet1!{names1['Deal ID']}:{names1['Deal ID']}, {names3['Deal ID']}2, "
            f"Sheet1!{names1['Best suggested quantity']}:"
            f"{names1['Best suggested quantity']})"
        ),
        "SSQ": (
            f"=SUMIF(Sheet1!{names1['Deal ID']}:{names1['Deal ID']}, {names3['Deal ID']}2, "
            f"Sheet1!{names1['System Suggested Quantity']}:"
            f"{names1['System Suggested Quantity']})"
        ),
    }

    apply_formulas(sheet1, formulas_sheet1, names1)
    apply_formulas(sheet3, formulas_sheet3, names3)

    autofill_formulas(sheet1, formulas_sheet1, names1)
    autofill_formulas(sheet3, formulas_sheet3, names3)

    hide_columns_range(
        sheet1,
        names1["Effectiveness"],
        names1["Empty ADS Check Profit"],
    )

    color_entire_column(sheet1, names1["Best suggested quantity"], (146, 208, 80))

    book.save(filename)

    if echo:
        print("Formulas filled successfully (openpyxl)")


if __name__ == "__main__":
    file_name = "../../output/13.05.2024/test - Copy.xlsx"
    main(file_name)
